import os

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def add_row_to_excel(file_path, sheet_name, row_data):
    """Создаём файл excel если его нет, добавляем данные в конец файла"""
    if os.path.exists(file_path):
        workbook = load_workbook(filename=file_path)
    else:
        workbook = Workbook()
        workbook.create_sheet(title=sheet_name)

    sheet = workbook[sheet_name]
    sheet.append(row_data)
    workbook.save(file_path)
